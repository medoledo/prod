#accounts/views.py

from django.conf import settings
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, BasePermission
from .serializers import (
    TeacherProfileSerializer,
    StudentProfileSerializer,
    UserSerializer,
    CenterSerializer,
    GradeSerializer,
    SubjectSerializer,
    AssistantProfileSerializer
)
from .models import (
    User,
    StudentProfile,
    TeacherProfile,
    AssistantProfile,
    Subject,
    Grade,
    Center,
    Payment
)
from django.db.models import Count, Q, Sum, Case, When, IntegerField
from rest_framework_simplejwt.views import TokenObtainPairView , TokenRefreshView
from .serializers import CustomTokenObtainPairSerializer , CustomTokenRefreshSerializer
from .permissions import (
    IsTeacher, IsStudent, IsAssistant, IsAdmin,
    IsTeacherOrAdmin, IsTeacherOrAssistant, IsTeacherAssistantOrAdmin
)
from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.utils import timezone
from datetime import timedelta
from session.models import Session, SessionAttendance, SessionTestScore
from studymaterials.models import StudyWeek
from quizzes.models import Quiz, QuizSubmission, QuizCenter

class PublicKeyView(APIView):
    permission_classes = []

    def get(self, request):
        # The key is already loaded into memory by Django settings.
        # This avoids inefficient file I/O on every request and hardcoded paths.
        public_key = settings.SIMPLE_JWT.get('VERIFYING_KEY')
        if not public_key:
            return Response({'error': 'Public key not configured on the server.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({'public_key': public_key})

# Authentication Views
class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class CustomTokenRefreshView(APIView):
    permission_classes = []

    def post(self, request):
        serializer = CustomTokenRefreshSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.validated_data)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_dashboard(request):
    if request.user.role != 'admin':
        return Response({'detail': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)

    # Use distinct counts to avoid join multiplication issues
    teachers = TeacherProfile.objects.annotate(
        assistants_count=Count('assistants', distinct=True),
        active_students=Count(
            'students',
            filter=Q(students__is_approved=True),
            distinct=True
        ),
        inactive_students=Count(
            'students',
            filter=Q(students__is_approved=False),
            distinct=True
        )
    )

    stats = {
        'teachers_count': TeacherProfile.objects.count(),
        'total_active_students': StudentProfile.objects.filter(is_approved=True).count(),
        'total_inactive_students': StudentProfile.objects.filter(is_approved=False).count(),
        'assistants_count': AssistantProfile.objects.count(),
        'teachers': list(teachers.values(
            'id',
            'full_name',
            'assistants_count',
            'active_students',
            'inactive_students'
        )),
        'recent_payments': list(Payment.objects.all().order_by('-date')[:5].values(
            'teacher__full_name', 'amount', 'date'
        ))
    }
    return Response(stats, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsTeacher])
def teacher_dashboard(request):
    teacher = request.user.teacher_profile
    stats = {
        'students_count': StudentProfile.objects.filter(teacher=teacher).count(),
        'assistants_count': AssistantProfile.objects.filter(teacher=teacher).count(),
        'centers': list(Center.objects.filter(teacher=teacher).values('id', 'name')),
        'pending_approvals': StudentProfile.objects.filter(teacher=teacher, is_approved=False).count()
    }
    return Response(stats, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsStudent])
def student_dashboard(request):
    try:
        student = request.user.student_profile
        teacher = student.teacher
        now = timezone.now()

        # 1. Profile Data
        profile_data = StudentProfileSerializer(student).data
        profile_data['subject_name'] = teacher.subject.name if teacher.subject else None

        # 2. Recent Session Test Scores (last 3)
        recent_session_scores = SessionTestScore.objects.filter(
            student=student
        ).select_related('session').order_by('-created_at')[:3]
        session_scores_data = [
            {
                'session_title': score.session.title,
                'score': f"{score.score}/{score.session.test_max_score}",
                'date': score.session.date
            }
            for score in recent_session_scores if score.session.test_max_score
        ]

        # 3. Attendance Statistics
        total_assigned_sessions = Session.objects.filter(
            grade=student.grade,
            center=student.center,
            date__lte=now.date()
        ).count()

        attendance_records = SessionAttendance.objects.filter(
            student=student,
            attended=True
        ).select_related('session__center')

        attended_from_own_center = attendance_records.filter(session__center=student.center).count()
        attended_from_other_center = attendance_records.exclude(session__center=student.center).count()
        total_attendance = attended_from_own_center + attended_from_other_center
        attendance_percentage = (total_attendance / total_assigned_sessions * 100) if total_assigned_sessions > 0 else 0

        attendance_data = {
            'total_assigned_sessions': total_assigned_sessions,
            'attended_from_own_center': attended_from_own_center,
            'attended_from_other_center': attended_from_other_center,
            'total_attendance': total_attendance,
            'percentage': f"{attendance_percentage:.0f}%"
        }

        # 4. Recent Study Weeks (last 3)
        recent_study_weeks = StudyWeek.objects.filter(
            grade=student.grade,
            centers=student.center
        ).order_by('-date_created')[:3]
        study_weeks_data = [{'id': week.id, 'title': week.title, 'date_created': week.date_created} for week in recent_study_weeks]

        # 5. Recent Online Quizzes (last 3)
        recent_quizzes = Quiz.objects.filter(
            grade=student.grade,
            quizcenter__center=student.center
        ).annotate(total_points=Sum('questions__points')).select_related('settings').order_by('-created_at').distinct()[:3]

        quiz_ids = [q.id for q in recent_quizzes]
        submissions_map = {s.quiz_id: s for s in QuizSubmission.objects.filter(quiz_id__in=quiz_ids, student=student, is_submitted=True)}
        quiz_centers_map = {qc.quiz_id: qc for qc in QuizCenter.objects.filter(quiz_id__in=quiz_ids, center=student.center)}

        quizzes_data = []
        for quiz in recent_quizzes:
            submission = submissions_map.get(quiz.id)
            quiz_center = quiz_centers_map.get(quiz.id)
            score_display = None

            if submission and quiz_center:
                settings = quiz.settings
                effective_release_time = quiz_center.close_date
                try:
                    if settings.timer_minutes > 0:
                        effective_release_time += timedelta(minutes=settings.timer_minutes)
                    is_quiz_closed = now > effective_release_time
                except OverflowError:
                    # If timer is too large, it's not closed yet.
                    is_quiz_closed = False
                    
                score_is_visible = (settings.score_visibility == 'immediate' or
                                    (settings.score_visibility == 'after_close' and is_quiz_closed) or
                                    (settings.score_visibility == 'manual' and submission.is_score_released))

                if score_is_visible:
                    total_points = quiz.total_points or 0
                    score_display = f"{submission.score:.1f} / {float(total_points):.1f}"
            
            quizzes_data.append({'id': quiz.id, 'title': quiz.title, 'score': score_display})

        # --- Constructing the final response ---
        final_data = {
            'profile': profile_data,
            'recent_session_scores': session_scores_data,
            'attendance_summary': attendance_data,
            'recent_study_weeks': study_weeks_data,
            'recent_online_quizzes': quizzes_data,
        }
        return Response(final_data, status=status.HTTP_200_OK)
    except StudentProfile.DoesNotExist:
        return Response({'error': 'Student profile not found.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': 'Could not retrieve dashboard data.', 'details': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAssistant])
def assistant_dashboard(request):
    assistant = request.user.assistant_profile
    stats = {
        'teacher': assistant.teacher.full_name,
        'students_count': StudentProfile.objects.filter(teacher=assistant.teacher).count(),
        'centers': list(Center.objects.filter(teacher=assistant.teacher).values('id', 'name'))
    }
    return Response(stats, status=status.HTTP_200_OK)

# Student Management
@api_view(['POST'])
@permission_classes([IsTeacherAssistantOrAdmin])  # Updated permission
def create_student(request):
    # 1. Create User
    user_data = {
        'username': request.data.get('username'),
        'password': request.data.get('password'),
        'email': request.data.get('email', ''),
        'role': 'student'
    }

    user_serializer = UserSerializer(data=user_data)
    if not user_serializer.is_valid():
        return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = user_serializer.save()
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # 2. Determine Teacher - now handles assistants
    try:
        if request.user.role == 'teacher':
            teacher = request.user.teacher_profile
        elif request.user.role == 'assistant':
            # Assistant must have a teacher profile
            if not hasattr(request.user, 'assistant_profile'):
                user.delete()
                return Response({'error': 'Assistant profile not found'}, status=400)
            teacher = request.user.assistant_profile.teacher
        else:  # admin
            if 'teacher' not in request.data:
                user.delete()
                return Response({'error': 'Teacher ID required for admin'}, status=400)
            teacher = TeacherProfile.objects.get(id=request.data['teacher'])
    except TeacherProfile.DoesNotExist:
        user.delete()
        return Response({'error': 'Invalid teacher ID'}, status=400)

    # 3. Prepare Profile Data
    profile_data = {
        'full_name': request.data.get('full_name'),
        'phone_number': request.data.get('phone_number'),
        'parent_number': request.data.get('parent_number'),
        'gender': request.data.get('gender'),
        'grade': request.data.get('grade'),
        'center': request.data.get('center')
    }

    # 4. Validate and Create Student Profile
    serializer = StudentProfileSerializer(
        data=profile_data,
        context={'request': request, 'teacher': teacher}
    )

    if serializer.is_valid():
        try:
            # Explicitly set user and teacher relationships
            student = serializer.save(
                user=user,
                teacher=teacher
            )
            return Response(StudentProfileSerializer(student).data, status=201)
        except Exception as e:
            user.delete()
            return Response({'error': str(e)}, status=400)

    # Cleanup if validation fails
    user.delete()
    return Response(serializer.errors, status=400)


@api_view(['POST'])
@permission_classes([IsAdmin])
def approve_students(request):
    # Validate request data
    student_ids = request.data.get('student_ids', [])
    is_approved = request.data.get('is_approved', None)

    if not isinstance(student_ids, list):
        return Response({'detail': 'student_ids should be a list'}, status=status.HTTP_400_BAD_REQUEST)

    if not isinstance(is_approved, bool):
        return Response({'detail': 'is_approved boolean field required'}, status=status.HTTP_400_BAD_REQUEST)

    # Process students
    updated = []
    not_found = []

    for student_id in student_ids:
        try:
            student = StudentProfile.objects.get(id=student_id)
            if student.is_approved != is_approved:
                student.is_approved = is_approved
                student.save()
            updated.append(student_id)
        except StudentProfile.DoesNotExist:
            not_found.append(student_id)

    response_data = {
        'message': f'Successfully updated {len(updated)} students',
        'is_approved': is_approved,
        'count_updated': len(updated),
        'count_not_found': len(not_found),
        'updated_students': updated,
        'not_found_ids': not_found
    }

    status_code = status.HTTP_200_OK if not not_found else status.HTTP_207_MULTI_STATUS
    return Response(response_data, status=status_code)

@api_view(['GET'])
@permission_classes([IsTeacherAssistantOrAdmin])
def list_students(request):
    # Base queryset
    if request.user.role == 'admin':
        queryset = StudentProfile.objects.all()
    else:
        # For teachers and assistants, get their associated teacher
        teacher = (
            request.user.teacher_profile
            if request.user.role == 'teacher'
            else request.user.assistant_profile.teacher
        )
        queryset = StudentProfile.objects.filter(teacher=teacher)

    # Apply filters for all roles
    search_query = request.GET.get('search')
    center_id = request.GET.get('center_id')
    grade_id = request.GET.get('grade_id')

    if search_query:
        queryset = queryset.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(parent_number__icontains=search_query)
        )

    if center_id:
        queryset = queryset.filter(center_id=center_id)

    if grade_id:
        queryset = queryset.filter(grade_id=grade_id)

    serializer = StudentProfileSerializer(queryset, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsTeacherAssistantOrAdmin])
def student_detail(request, pk):
    # 1) Fetch student with permission scope, and determine the 'teacher' for context
    try:
        if request.user.role == 'admin':
            student = StudentProfile.objects.get(pk=pk)
            teacher_obj = student.teacher
        elif request.user.role == 'teacher':
            teacher_obj = request.user.teacher_profile
            student = StudentProfile.objects.get(pk=pk, teacher=teacher_obj)
        else:  # assistant
            teacher_obj = request.user.assistant_profile.teacher
            student = StudentProfile.objects.get(pk=pk, teacher=teacher_obj)
    except StudentProfile.DoesNotExist:
        return Response(
            {'detail': 'Student not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    # Prepare serializer context
    ctx = {
        'request': request,
        'teacher': teacher_obj
    }

    # 2) GET: include `username` in the representation
    if request.method == 'GET':
        serializer = StudentProfileSerializer(student, context=ctx)
        return Response(serializer.data)

    # 3) PUT: allow partial updates to profile _and_ to the related User
    elif request.method == 'PUT':
        serializer = StudentProfileSerializer(
            student,
            data=request.data,
            partial=True,
            context=ctx
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # 4) DELETE: remove both profile and user account
    else:  # DELETE
        user = student.user
        student.delete()
        user.delete()
        return Response(
            {'detail': 'Student deleted successfully'},
            status=status.HTTP_204_NO_CONTENT
        )
# Add to views.py
@api_view(['GET'])
@permission_classes([IsAdmin])
def teacher_students(request, teacher_id):
    """
    Get all students for a specific teacher
    """
    try:
        teacher = TeacherProfile.objects.get(pk=teacher_id)
        students = StudentProfile.objects.filter(teacher=teacher)
        serializer = StudentProfileSerializer(students, many=True)
        return Response(serializer.data)
    except TeacherProfile.DoesNotExist:
        return Response({'detail': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAdmin])
def teacher_active_students(request, teacher_id):
    """
    Get active students for a specific teacher
    """
    try:
        teacher = TeacherProfile.objects.get(pk=teacher_id)
        students = StudentProfile.objects.filter(teacher=teacher, is_approved=True)
        serializer = StudentProfileSerializer(students, many=True)
        return Response(serializer.data)
    except TeacherProfile.DoesNotExist:
        return Response({'detail': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAdmin])
def teacher_inactive_students(request, teacher_id):
    """
    Get inactive students for a specific teacher
    """
    try:
        teacher = TeacherProfile.objects.get(pk=teacher_id)
        students = StudentProfile.objects.filter(teacher=teacher, is_approved=False)
        serializer = StudentProfileSerializer(students, many=True)
        return Response(serializer.data)
    except TeacherProfile.DoesNotExist:
        return Response({'detail': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)





# Teacher Management
@api_view(['POST'])
@permission_classes([IsAdmin])
def create_teacher_profile(request):
    # 1. Create User
    user_data = {
        'username': request.data.get('username'),
        'password': request.data.get('password'),
        'email': request.data.get('email', ''),
        'role': 'teacher'
    }

    user_serializer = UserSerializer(data=user_data)
    if not user_serializer.is_valid():
        return Response(user_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = user_serializer.save()
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # 2. Create Teacher Profile
    profile_data = {
        'full_name': request.data.get('full_name'),
        'phone_number': request.data.get('phone_number'),
        'gender': request.data.get('gender'),
        'brand': request.data.get('brand'),
        'subject': request.data.get('subject'),
        'grades': request.data.get('grades', [])
    }

    # 3. Validate and Create Profile
    serializer = TeacherProfileSerializer(
        data=profile_data,
        context={'request': request}
    )

    if serializer.is_valid():
        try:
            # Explicitly set the user relationship
            teacher = serializer.save(user=user)
            return Response({
                'user': UserSerializer(user).data,
                'profile': TeacherProfileSerializer(teacher).data
            }, status=status.HTTP_201_CREATED)
        except Exception as e:
            user.delete()
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # Cleanup if validation fails
    user.delete()
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsTeacherOrAdmin])
def my_teacher_profile(request):
    try:
        profile = request.user.teacher_profile
        serializer = TeacherProfileSerializer(profile)
        return Response(serializer.data)
    except TeacherProfile.DoesNotExist:
        return Response({'detail': 'Profile not found'}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([IsAdmin])
def list_teachers(request):
    teachers = TeacherProfile.objects.prefetch_related(
        'students',
        'students__grade',
        'students__center',
        'subject'
    ).all()
    serializer = TeacherProfileSerializer(teachers, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAdmin])
def teacher_detail(request, pk):
    try:
        teacher = TeacherProfile.objects.get(pk=pk)
    except TeacherProfile.DoesNotExist:
        return Response({'detail': 'Teacher not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = TeacherProfileSerializer(teacher)
        return Response(serializer.data)

    elif request.method == 'PUT':
        # Handle password update first
        if 'password' in request.data:
            user = teacher.user
            new_password = request.data.pop('password')  # Remove password from request data
            user.set_password(new_password)
            user.save()

        # Process regular fields
        serializer = TeacherProfileSerializer(teacher, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        # To ensure a clean deletion, we must remove all dependent objects first.
        # This comprehensive approach handles potential PROTECT relations on models
        # like Session, StudyWeek, StudentProfile, etc.

        # Get all centers for the teacher, which are a nexus for other data.
        teacher_centers = Center.objects.filter(teacher=teacher)
        teacher_center_ids = teacher_centers.values_list('id', flat=True)

        # 1. Delete Sessions (and by cascade, their attendance/scores).
        Session.objects.filter(center_id__in=teacher_center_ids).delete()

        # 2. Delete Study Weeks (and by cascade, their materials).
        StudyWeek.objects.filter(centers__id__in=teacher_center_ids).delete()

        # 3. Delete all related students and their user accounts.
        User.objects.filter(student_profile__teacher=teacher).delete()

        # 4. Delete all related assistants and their user accounts.
        User.objects.filter(assistant_profile__teacher=teacher).delete()

        # 5. Delete all related quizzes.
        Quiz.objects.filter(teacher=teacher).delete()

        # 6. Delete all related centers.
        teacher_centers.delete()

        # 7. Finally, delete the teacher's user account, which cascades to the TeacherProfile.
        teacher.user.delete()

        return Response(
            {'detail': 'Teacher and all related data deleted successfully'},
            status=status.HTTP_204_NO_CONTENT
        )

# Assistant Management (consistent with student API)
@api_view(['GET'])
@permission_classes([IsTeacherOrAdmin])
def assistant_list(request):
    # Base queryset
    if request.user.role == 'admin':
        queryset = AssistantProfile.objects.all()
    else:
        queryset = AssistantProfile.objects.filter(
            teacher=request.user.teacher_profile
        )

    # Apply filters
    search_query = request.GET.get('search')
    teacher_id = request.GET.get('teacher_id')

    if search_query:
        queryset = queryset.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )

    if teacher_id:
        queryset = queryset.filter(teacher_id=teacher_id)

    serializer = AssistantProfileSerializer(queryset, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsTeacherOrAdmin])
def create_assistant(request):
    data = request.data.copy()

    # For teachers: automatically set teacher to themselves
    if request.user.role == 'teacher':
        data['teacher'] = request.user.teacher_profile.id

    serializer = AssistantProfileSerializer(
        data=data,
        context={'request': request}
    )

    if serializer.is_valid():
        # Additional validation for admin users
        if request.user.role == 'admin' and 'teacher' not in data:
            return Response(
                {"teacher": "Teacher field is required for admin users"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            assistant = serializer.save()
            return Response(
                AssistantProfileSerializer(assistant).data,
                status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response({'error': str(e)}, status=400)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsTeacherOrAdmin])
def assistant_detail(request, pk):
    try:
        if request.user.role == 'admin':
            assistant = AssistantProfile.objects.get(pk=pk)
        else:
            assistant = AssistantProfile.objects.get(
                pk=pk,
                teacher=request.user.teacher_profile
            )
    except AssistantProfile.DoesNotExist:
        return Response(
            {'detail': 'Assistant not found'},
            status=status.HTTP_404_NOT_FOUND
        )

    if request.method == 'GET':
        serializer = AssistantProfileSerializer(assistant)
        return Response(serializer.data)

    elif request.method == 'PUT':
        try:
            # Make a mutable copy of the request data
            data = request.data.copy()

            # For teachers: prevent changing teacher association
            if request.user.role == 'teacher':
                if 'teacher' in data:
                    # Teachers can't change teacher association
                    if int(data['teacher']) != request.user.teacher_profile.id:
                        return Response(
                            {'teacher': 'Cannot change teacher association'},
                            status=status.HTTP_403_FORBIDDEN
                        )
                else:
                    # Ensure teacher remains the same
                    data['teacher'] = request.user.teacher_profile.id

            serializer = AssistantProfileSerializer(
                assistant,
                data=data,
                partial=True,
                context={'request': request}
            )

            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response(
                {'error': f'Error updating assistant: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    elif request.method == 'DELETE':
        try:
            # Get user and delete both assistant profile and user account
            user = assistant.user
            assistant.delete()
            user.delete()
            return Response(
                {'detail': 'Assistant deleted successfully'},
                status=status.HTTP_204_NO_CONTENT
            )
        except Exception as e:
            return Response(
                {'error': f'Error deleting assistant: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# Center Management
@api_view(['POST'])
@permission_classes([IsTeacherOrAdmin])
def create_center(request):
    # Make a mutable copy of the request data
    data = request.data.copy()

    # Handle teacher role: auto-assign their ID
    if request.user.role == 'teacher':
        data['teacher'] = request.user.teacher_profile.id

    # Handle admin role: validate teacher ID
    elif request.user.role == 'admin' and 'teacher' not in data:
        return Response(
            {'teacher': 'Teacher ID is required for admin users'},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = CenterSerializer(data=data, context={'request': request})

    if serializer.is_valid():
        try:
            # Check for existing center with same name for this teacher
            teacher_id = data['teacher']
            center_name = data['name']

            if Center.objects.filter(
                teacher_id=teacher_id,
                name__iexact=center_name
            ).exists():
                return Response(
                    {'detail': 'A center with this name already exists for the teacher'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Save if validation passes
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except KeyError as e:
            return Response(
                {'detail': f'Missing required field: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_centers(request):
    # Role-based filtering
    if request.user.role == 'admin':
        centers = Center.objects.all()
    elif request.user.role == 'teacher':
        centers = Center.objects.filter(teacher=request.user.teacher_profile)
    elif request.user.role == 'assistant':
        # Get centers from the assistant's teacher
        centers = Center.objects.filter(teacher=request.user.assistant_profile.teacher)
    elif request.user.role == 'student':
        # Get only the student's assigned center
        centers = Center.objects.filter(id=request.user.student_profile.center_id)
    else:
        return Response(
            {'detail': 'Invalid user role'},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = CenterSerializer(centers, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)




@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def grade_list(request):
    if request.method == 'GET':
        # Handle different user roles
        if request.user.role == 'admin':
            grades = Grade.objects.all()
        elif request.user.role == 'teacher':
            # Get grades assigned to this teacher
            grades = request.user.teacher_profile.grades.all()
        elif request.user.role == 'assistant':
            # Get grades from the assistant's teacher
            grades = request.user.assistant_profile.teacher.grades.all()
        elif request.user.role == 'student':
            # Get only the student's assigned grade
            student_profile = request.user.student_profile
            grades = Grade.objects.filter(id=student_profile.grade.id)
        else:
            return Response(
                {'detail': 'Invalid user role'},
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = GradeSerializer(grades, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        # POST remains admin-only as before
        if request.user.role != 'admin':
            return Response(
                {'detail': 'Only admins can create grades'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = GradeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])  # Base permission for all
def subject_list(request):
    if request.method == 'GET':
        subjects = Subject.objects.all()
        serializer = SubjectSerializer(subjects, many=True)
        return Response(serializer.data)

    elif request.method == 'POST':
        # Explicit admin check for POST
        if request.user.role != 'admin':
            return Response({'detail': 'Only admins can create subjects'},
                          status=status.HTTP_403_FORBIDDEN)

        serializer = SubjectSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)





@api_view(['GET'])
@permission_classes([IsTeacher])
def export_students_to_excel(request):
    teacher = request.user.teacher_profile
    students = StudentProfile.objects.filter(teacher=teacher).select_related(
        'grade', 'center', 'user'
    )

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Column headers
    headers = [
        'Student ID', 'Username', 'Full Name', 'Phone Number',
        'Parent Number', 'Gender', 'Grade', 'Center',
        'Approval Status', 'Added By'
    ]

    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write student data
    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1, value=student.student_id)
        ws.cell(row=row_num, column=2, value=student.user.username)
        ws.cell(row=row_num, column=3, value=student.full_name)
        ws.cell(row=row_num, column=4, value=student.phone_number)
        ws.cell(row=row_num, column=5, value=student.parent_number)
        ws.cell(row=row_num, column=6, value=student.get_gender_display())
        ws.cell(row=row_num, column=7, value=student.grade.name)
        ws.cell(row=row_num, column=8, value=student.center.name)
        ws.cell(row=row_num, column=9, value="Active" if student.is_approved else "Inactive")
        ws.cell(row=row_num, column=10, value=student.added_by)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{teacher.full_name}_students.xlsx"'
    )

    wb.save(response)
    return response